import openpyxl

PATH = '/home/ctx/minjoon/'
txt_file_name = 'music.txt'
excel_file_name = 'music.xlsx'

txt_file = open(PATH + txt_file_name)
excel_file = openpyxl.load_workbook(PATH + excel_file_name)

sheet = excel_file['Sheet1']

lines = txt_file.readlines()

row = 1
for idx, line in enumerate(lines):
    if '재생 담기' in line:
        title = lines[idx+1].replace('\n', '')
        artist = lines[idx+3].replace('\n', '')
        album = lines[idx+4].replace('\n', '')
        
        title_cell = sheet.cell(row = row, column = 1)
        artist_cell = sheet.cell(row = row, column = 2)
        album_cell = sheet.cell(row = row, column = 3)
            
        title_cell.value = title
        artist_cell.value = artist
        album_cell.value = album
        row += 1
        
        print(title)

excel_file.save(PATH + excel_file_name)

excel_file.close()
txt_file.close()